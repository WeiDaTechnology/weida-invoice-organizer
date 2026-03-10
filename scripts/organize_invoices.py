#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
发票整理主脚本
功能：
1. 扫描文件夹中的 PDF 发票
2. 提取发票信息
3. 复制到新文件夹并重命名
4. 导出到 Excel

用法：
    python organize_invoices.py <folder_path> [output_folder_name]
    
示例：
    python organize_invoices.py "C:/invoices"              # 默认输出到 已整理 文件夹
    python organize_invoices.py "C:/invoices" "已报销"     # 输出到指定文件夹
"""

import sys
import shutil
from pathlib import Path
from typing import Optional
from openpyxl import Workbook, load_workbook

# 导入提取脚本
from extract_invoice_info import extract_invoice_info

# 默认模板路径（技能内部）
DEFAULT_TEMPLATE = Path(__file__).parent.parent / "templates" / "费用报销清单明细表-demo.xlsx"


def is_generated_output_file(root_folder: Path, pdf_path: Path, output_folder_name: str) -> bool:
    """
    排除历史整理结果，避免重复运行时把已整理文件再次当作输入。
    """
    try:
        relative_parts = pdf_path.relative_to(root_folder).parts
    except ValueError:
        return False

    if not relative_parts:
        return False

    top_level_dir = relative_parts[0]
    return top_level_dir == output_folder_name or top_level_dir.startswith("已整理")


def process_invoices(folder_path: str, output_folder_name: str = "已整理"):
    """
    处理文件夹中的所有发票 PDF
    
    Args:
        folder_path: 发票 PDF 所在文件夹
        output_folder_name: 输出文件夹名称（默认为"已整理"）
    
    Returns:
        dict: 处理结果统计
    """
    folder = Path(folder_path)
    if not folder.exists():
        return {"error": f"文件夹不存在：{folder_path}"}
    
    # 在源文件夹下创建输出文件夹
    output = folder / output_folder_name
    output.mkdir(parents=True, exist_ok=True)
    
    print(f"输出文件夹：{output}")
    
    # 递归找出所有 PDF 文件，方便直接整理包含子文件夹的报销目录。
    pdf_files = [
        f for f in folder.rglob("*.pdf")
        if not is_generated_output_file(folder, f, output_folder_name)
    ]
    if not pdf_files:
        return {"error": f"未找到 PDF 文件：{folder_path}"}
    
    print(f"找到 {len(pdf_files)} 个 PDF 文件")
    
    # 存储所有发票信息
    invoices = []
    stats = {
        "total": len(pdf_files),
        "success": 0,
        "failed": 0,
        "copied": 0,
        "errors": []
    }
    
    for pdf_path in pdf_files:
        print(f"\n处理：{pdf_path.name}")
        
        # 提取信息
        info = extract_invoice_info(str(pdf_path))
        
        if info.get("error"):
            stats["failed"] += 1
            stats["errors"].append({
                "file": pdf_path.name,
                "error": info["error"]
            })
            print(f"  ❌ 提取失败：{info['error']}")
            continue
        
        # 验证必要字段
        if not info.get("invoice_number") or not info.get("date"):
            stats["failed"] += 1
            stats["errors"].append({
                "file": pdf_path.name,
                "error": "缺少必要字段（发票号码或日期）"
            })
            print(f"  ❌ 缺少必要字段")
            continue
        
        # 生成新文件名：金额_摘要_发票号码.pdf
        amount_str = f"{info['amount']:.2f}".replace('.', '_')
        summary = info.get('item_name') or info.get('seller_name', '未知')
        summary = sanitize_filename_part(summary, max_length=20)
        
        new_name = f"{amount_str}_{summary}_{info['invoice_number']}.pdf"
        new_path = output / new_name
        
        # 复制文件到输出文件夹并重命名
        try:
            shutil.copy2(pdf_path, new_path)
            stats["copied"] += 1
            print(f"  ✅ 已复制：{new_name}")
        except Exception as e:
            stats["failed"] += 1
            stats["errors"].append({
                "file": pdf_path.name,
                "error": f"复制失败：{str(e)}"
            })
            print(f"  ❌ 复制失败：{e}")
            continue
        
        stats["success"] += 1
        
        # 添加到列表
        invoices.append({
            "date": info["date"],
            "seller": info.get("seller_name", ""),
            "item": info.get("item_name", ""),
            "amount": info["amount"],
            "invoice_number": info["invoice_number"],
            "file": new_name
        })
    
    # 导出到 Excel（使用默认模板）
    if invoices:
        excel_output = output / "费用报销清单明细表.xlsx"
        if DEFAULT_TEMPLATE.exists():
            export_to_excel(invoices, str(DEFAULT_TEMPLATE), excel_output)
            print(f"\n✅ Excel 已导出：{excel_output}")
        else:
            export_to_excel(invoices, None, excel_output)
            print(f"\n✅ 未找到模板，已使用内置表头导出 Excel：{excel_output}")
    
    return stats


def sanitize_filename_part(value: str, max_length: int = 20) -> str:
    """
    清理文件名片段，避免 Windows 非法字符和多余空白。
    """
    if not value:
        return "未知"

    invalid_chars = '<>:"/\\|?*'
    sanitized = value
    for ch in invalid_chars:
        sanitized = sanitized.replace(ch, "_")

    sanitized = " ".join(sanitized.split()).strip(" .")
    return (sanitized or "未知")[:max_length]


def create_default_workbook():
    """
    创建一个可直接写入的默认报销清单工作簿。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "费用报销清单"
    ws.append(["编号", "时间", "用途（详细用途）", "金额", "责任人", "发票号", "摘要"])
    ws.append([None, None, "合计：", "=SUM(D2:D2)", None, None, None])
    return wb, ws


def export_to_excel(invoices: list, template_path: Optional[str], output_path: str):
    """
    将发票信息导出到 Excel 模板
    
    Excel 模板列：
    1. 编号
    2. 时间
    3. 用途（详细用途）
    4. 金额
    5. 责任人（留空）
    6. 发票号
    7. 摘要
    
    最后一行是统计行：合计：总金额（支持公式）
    """
    # 优先使用模板；缺失时自动创建默认工作簿，避免技能因模板文件缺失而失效。
    if template_path:
        wb = load_workbook(template_path)
        ws = wb.active
    else:
        wb, ws = create_default_workbook()
    
    # 找到数据开始行（跳过表头）
    start_row = 2
    
    # 检查模板最后一行是否有公式（保留公式并动态调整范围）
    has_formula = False
    formula_col = None
    formula_start_row = None
    label_cell = None
    
    # 检查模板最后一行第 4 列（金额列）是否有公式
    template_last_row = ws.max_row
    amount_cell = ws.cell(row=template_last_row, column=4)
    label_cell = ws.cell(row=template_last_row, column=3)
    
    if amount_cell.value and str(amount_cell.value).startswith('='):
        has_formula = True
        # 解析公式，提取列和起始行，例如 =SUM(D2:D14) -> 列 D, 起始行 2
        formula_str = str(amount_cell.value)
        import re
        match = re.search(r'=SUM\(([A-Z]+)(\d+):[A-Z]+\d+\)', formula_str, re.IGNORECASE)
        if match:
            formula_col = match.group(1).upper()  # 列字母，如 D
            formula_start_row = int(match.group(2))  # 起始行，如 2
            print(f"  📊 检测到模板公式：{formula_str} (列={formula_col}, 起始行={formula_start_row})")
        else:
            # 无法解析，退化为固定值
            has_formula = False
            print(f"  ⚠️ 公式格式无法解析，使用固定值：{formula_str}")
    
    # 清除数据行（从第 2 行开始到倒数第二行，保留最后一行模板行）
    for row in range(start_row, template_last_row):
        for col in range(1, 8):
            ws.cell(row=row, column=col).value = None
    
    # 填充数据
    for i, inv in enumerate(invoices, start=1):
        row = start_row + i - 1
        ws.cell(row=row, column=1, value=i)  # 编号
        ws.cell(row=row, column=2, value=inv["date"])  # 时间
        ws.cell(row=row, column=3, value=inv["seller"])  # 用途（销售方）
        ws.cell(row=row, column=4, value=inv["amount"])  # 金额
        ws.cell(row=row, column=5, value=None)  # 责任人（留空）
        ws.cell(row=row, column=6, value=inv["invoice_number"])  # 发票号
        ws.cell(row=row, column=7, value=inv["item"])  # 摘要
    
    # 处理统计行
    last_row = start_row + len(invoices)
    
    if has_formula and formula_col and formula_start_row:
        # 动态构建公式，根据实际数据行数调整结束行
        # 例如：=SUM(D2:D34)
        formula_end_row = last_row - 1  # 数据最后一行（统计行前一行的数据）
        dynamic_formula = f"=SUM({formula_col}{formula_start_row}:{formula_col}{formula_end_row})"
        ws.cell(row=last_row, column=3, value=label_cell.value if label_cell.value else "合计：")
        ws.cell(row=last_row, column=4, value=dynamic_formula)  # 使用动态公式
        print(f"  ✅ 使用动态公式：{dynamic_formula}")
    else:
        # 使用固定值（向后兼容）
        total_amount = sum(inv["amount"] for inv in invoices)
        ws.cell(row=last_row, column=3, value="合计：")
        ws.cell(row=last_row, column=4, value=round(total_amount, 2))
        print(f"  ✅ 使用固定值合计：{round(total_amount, 2)}")
    
    # 保存
    wb.save(output_path)


def main():
    if len(sys.argv) < 2:
        print("Usage: python organize_invoices.py <folder_path> [output_folder_name]")
        print("Example: python organize_invoices.py ./invoices")
        print("         python organize_invoices.py ./invoices 已报销")
        sys.exit(1)
    
    folder_path = sys.argv[1]
    output_folder_name = sys.argv[2] if len(sys.argv) > 2 else "已整理"
    
    result = process_invoices(folder_path, output_folder_name)

    if result.get("error"):
        print(f"\n❌ {result['error']}")
        sys.exit(1)
    
    print("\n" + "="*50)
    print("处理完成!")
    print(f"  总计：{result['total']}")
    print(f"  成功：{result['success']}")
    print(f"  失败：{result['failed']}")
    print(f"  已复制：{result['copied']}")
    
    if result.get("errors"):
        print("\n错误详情:")
        for err in result["errors"]:
            print(f"  - {err['file']}: {err['error']}")


if __name__ == "__main__":
    main()
